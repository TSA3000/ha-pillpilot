"""Build / refresh ``medicines_se.json`` from Läkemedelsverket's open-data export.

This is a build tool — runs locally, commits the resulting JSON.
Not shipped in the integration zip.

Source
------
Läkemedelsverket's "Färdig sökning med alla läkemedel" via Sveriges
dataportal (dataset 140_5467) → distribution `Lakemedelsprodukter.xlsx`.
Contains every medicine in Sök läkemedelsfakta — approved, registered,
temporarily withdrawn, and deregistered. Updated nightly. Free use
under Sweden's open-data law (öppna data-lagen, 2022:818).

  https://www.dataportal.se/datasets/140_5467

The export has one row per strength+form combination, not per medicine.
"Alvedon" appears as ~10 rows (250 mg suppositorium, 500 mg tablett,
24 mg/ml oral lösning, …). The script groups rows by ``Namn`` so the
output JSON has one entry per medicine name with combined forms.

Filters applied
---------------
- ``H/V == "HUM"`` — skip veterinary.
- ``Registrerings-status in ("Godkänd", "Registrerad")`` — skip
  Avregistrerad / Tillfälligt återkallad.
Everything human and currently on the market is kept regardless of
``Produktkategori`` (homeopathic, herbal, allergens included so the
search dropdown finds whatever the user actually takes).

Field mapping
-------------
Source column                         medicines_se.json
  Namn                              → name (group key)
  Verksamt ämne (förenklat)         → active_substance
  ATC-kod                           → atc_code
  Tidigare läkemedelsnamn           → aliases[] (split on commas; ≠ name)
  Styrka, Form, NPL-id              → variants[] (one per strength/form)

Variants array
--------------
Each medicine has a ``variants`` list — one entry per distinct
(strength, form) pair. ``Concerta`` has 4 variants (18/27/36/54 mg
Depottablett); ``Alvedon`` has ~11. The variant carries the precise
``npl_id``, ``strength``, and ``form`` strings, enabling per-variant
linking to Fass / Läkemedelsverket and smart strength/form pickers
in the panel. Parallel imports (multiple NPLs for the same
strength+form) are deduped — first NPL wins.

Existing curated aliases on a medicine that's already in the JSON are
preserved and merged with any newly-discovered Tidigare-läkemedelsnamn
values (curated wins on conflict — never overwritten).

Usage
-----
    pip install openpyxl    # only needed for .xlsx input
    python tools/build_medicines_se.py \\
        --input ~/Downloads/Lakemedelsprodukter.xlsx \\
        --output custom_components/pillpilot/medicines_se.json

Add ``--dry-run`` to print stats without touching the file.
Also accepts .csv, .tsv, .xml, .json — formats fall through to the
generic record loaders and use the same ``FIELD_ALIASES`` for header
detection.

The script bumps ``list_version`` to today's date with an ``-N``
suffix (incrementing if today already has a release).
"""
from __future__ import annotations

import argparse
import csv
import json
import re
import sys
import xml.etree.ElementTree as ET
from collections import defaultdict
from datetime import date
from pathlib import Path
from typing import Any, Iterable

# Source columns vary across Läkemedelsverket exports; accept all common
# spellings so the script doesn't break on a column rename. First match
# wins. Edit if the export gains a new column convention.
FIELD_ALIASES: dict[str, tuple[str, ...]] = {
    "name": ("namn", "läkemedelsnamn", "produktnamn", "name"),
    "active_substance": (
        "verksamt ämne (förenklat)", "verksamt ämne", "aktiv substans",
        "substans", "active substance", "active_substance",
    ),
    "atc_code": ("atc-kod", "atckod", "atc", "atc code", "atc_code"),
    "npl_id": ("npl-id", "nplid", "npl", "npl_id", "npl id"),
    "common_forms": (
        "form", "beredningsform", "läkemedelsform", "dosage form",
        "common_forms",
    ),
    "status": (
        "registrerings-status", "registreringsstatus", "status",
        "godkännandestatus", "approval status",
    ),
    "human_or_vet": ("h/v", "hum/vet", "humanvet"),
    "previous_name": (
        "tidigare läkemedelsnamn", "tidigare namn", "previous name",
    ),
}

# Status values we KEEP — everything else is filtered out.
KEPT_STATUSES = ("godkänd", "registrerad")
KEPT_HV = ("hum",)


def _norm(s: Any) -> str:
    """Lowercase + strip + collapse whitespace for header matching."""
    if s is None:
        return ""
    return re.sub(r"\s+", " ", str(s).strip().lower())


def _is_kept_status(status: str) -> bool:
    # Exact match on the normalized status — substring matching breaks
    # because "registrerad" is a substring of "avregistrerad", so a
    # naive `in` check would keep deregistered medicines.
    return _norm(status) in KEPT_STATUSES


def _is_kept_hv(hv: str) -> bool:
    s = _norm(hv)
    return any(token == s for token in KEPT_HV)


def _split_forms(raw: str) -> list[str]:
    """Form is usually a single value, occasionally comma/semicolon-listed."""
    if not raw:
        return []
    parts = re.split(r"[,;/]\s*", raw.strip())
    return [p for p in (p.strip() for p in parts) if p]


def _build_column_map(headers: list[str]) -> dict[str, str | None]:
    norm_headers = {_norm(h): h for h in headers}
    out: dict[str, str | None] = {}
    for field, candidates in FIELD_ALIASES.items():
        out[field] = None
        for cand in candidates:
            if cand in norm_headers:
                out[field] = norm_headers[cand]
                break
    return out


def load_records_xlsx(path: Path) -> list[dict[str, Any]]:
    try:
        import openpyxl  # type: ignore
    except ImportError:
        raise SystemExit(
            "openpyxl is required for .xlsx input. Install it with "
            "`pip install openpyxl`, or save the file as .csv first."
        )
    # Läkemedelsverket's export puts data in a sheet named after the
    # extraction date (e.g. "2026-05-10"). Other sheets are "Filter"
    # and "Information" — skip those.
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    data_sheet = None
    for sn in wb.sheetnames:
        if sn.lower() in ("filter", "information"):
            continue
        data_sheet = wb[sn]
        break
    if data_sheet is None:
        raise SystemExit(f"no data sheet found in {path}")
    rows = list(data_sheet.iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(h) if h is not None else "" for h in rows[0]]
    return [dict(zip(header, row)) for row in rows[1:]]


def load_records_csv(path: Path) -> list[dict[str, str]]:
    for enc in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            with path.open(encoding=enc, newline="") as f:
                sample = f.read(4096)
                f.seek(0)
                try:
                    dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
                except csv.Error:
                    dialect = csv.excel
                reader = csv.DictReader(f, dialect=dialect)
                return list(reader)
        except UnicodeDecodeError:
            continue
    raise SystemExit(f"could not decode {path} with any common encoding")


def load_records_xml(path: Path) -> list[dict[str, str]]:
    tree = ET.parse(path)
    root = tree.getroot()

    def child_text(el: ET.Element, tag_lower: str) -> str:
        for c in el:
            if _norm(c.tag.split("}")[-1]) == tag_lower:
                return (c.text or "").strip()
        return ""

    records: list[dict[str, str]] = []
    for el in root.iter():
        name = ""
        for cand in FIELD_ALIASES["name"]:
            name = child_text(el, cand)
            if name:
                break
        if not name:
            continue
        rec: dict[str, str] = {}
        for field, candidates in FIELD_ALIASES.items():
            for cand in candidates:
                v = child_text(el, cand)
                if v:
                    rec[field] = v
                    break
        records.append(rec)
    return records


def load_records(path: Path) -> list[dict[str, Any]]:
    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        return load_records_xlsx(path)
    if suffix in (".csv", ".tsv"):
        return load_records_csv(path)
    if suffix == ".xml":
        return load_records_xml(path)
    if suffix == ".json":
        data = json.loads(path.read_text(encoding="utf-8"))
        if isinstance(data, list):
            return data
        if isinstance(data, dict) and "medicines" in data:
            return data["medicines"]
        raise SystemExit(f"unrecognized JSON shape in {path}")
    raise SystemExit(f"unsupported input format: {suffix}")


def _row_get(row: dict[str, Any], col_map: dict[str, str | None],
             field: str) -> str:
    """Read a field from a row using the resolved column map."""
    col = col_map.get(field)
    if col and col in row:
        v = row[col]
        if v is None:
            return ""
        return str(v).strip()
    return ""


def group_and_normalize(
    raw_rows: list[dict[str, Any]],
) -> tuple[list[dict[str, Any]], dict[str, int]]:
    """Filter raw rows and group by Namn into one entry per medicine.

    Each entry has a ``variants`` list of ``{npl_id, strength, form}`` —
    one per distinct (strength, form) for the brand. Parallel imports
    (multiple NPLs for the same user-facing strength+form combo) are
    deduped: the first NPL wins.
    """
    if not raw_rows:
        return [], {}
    headers = list(raw_rows[0].keys())
    cmap = _build_column_map(headers)
    if not cmap.get("name"):
        raise SystemExit(
            f"could not find a 'name' column in input. "
            f"Saw headers: {headers[:8]}..."
        )

    stats = {
        "total_rows": len(raw_rows),
        "skipped_no_name": 0,
        "skipped_status": 0,
        "skipped_vet": 0,
        "kept_rows": 0,
        "dedup_parallel_imports": 0,
    }

    grouped: dict[str, dict[str, Any]] = defaultdict(
        lambda: {
            "name": "", "atc": "", "substance": "",
            "aliases": set(),
            # Variant key (strength_norm, form_norm) → variant dict.
            # Dedupes parallel imports while preserving the first NPL.
            "variants": {},
        }
    )

    for row in raw_rows:
        name = _row_get(row, cmap, "name")
        if not name:
            stats["skipped_no_name"] += 1
            continue
        status = _row_get(row, cmap, "status")
        if status and not _is_kept_status(status):
            stats["skipped_status"] += 1
            continue
        hv = _row_get(row, cmap, "human_or_vet")
        if hv and not _is_kept_hv(hv):
            stats["skipped_vet"] += 1
            continue
        stats["kept_rows"] += 1

        key = _norm(name)
        g = grouped[key]
        if not g["name"]:
            g["name"] = name
        if not g["substance"]:
            g["substance"] = _row_get(row, cmap, "active_substance")
        if not g["atc"]:
            g["atc"] = _row_get(row, cmap, "atc_code")
        prev_raw = _row_get(row, cmap, "previous_name")
        if prev_raw:
            for prev in re.split(r"[,;]\s*", prev_raw):
                prev = prev.strip()
                if not prev or _norm(prev) == key:
                    continue
                g["aliases"].add(prev)

        strength = _row_get(row, cmap, "common_forms")  # placeholder
        # Variant key is (normalized strength, normalized form) so
        # parallel imports collapse. NPL of the first row wins.
        strength = _row_get_strength(row, cmap)
        form = _row_get(row, cmap, "common_forms")
        npl = _row_get(row, cmap, "npl_id")
        v_key = (_norm(strength), _norm(form))
        if v_key in g["variants"]:
            stats["dedup_parallel_imports"] += 1
            continue
        variant: dict[str, str] = {}
        if npl:
            variant["npl_id"] = npl
        if strength:
            variant["strength"] = strength
        if form:
            variant["form"] = form
        if variant:
            g["variants"][v_key] = variant

    out: list[dict[str, Any]] = []
    for g in grouped.values():
        rec: dict[str, Any] = {
            "name": g["name"],
            "aliases": sorted(g["aliases"], key=str.lower),
        }
        if g["substance"]:
            rec["active_substance"] = g["substance"]
        if g["atc"]:
            rec["atc_code"] = g["atc"]
        # Sort variants by (strength, form) for stable diff
        rec["variants"] = sorted(
            g["variants"].values(),
            key=lambda v: (
                _strength_sort_key(v.get("strength", "")),
                v.get("form", "").lower(),
            ),
        )
        out.append(rec)
    out.sort(key=lambda m: _norm(m["name"]))
    stats["unique_medicines"] = len(out)
    stats["total_variants"] = sum(len(m["variants"]) for m in out)
    return out, stats


def _row_get_strength(row: dict[str, Any], cmap: dict[str, str | None]) -> str:
    """Read the Styrka column. Listed separately because there's no
    schema-level 'strength' field — Läkemedelsverket exports use 'Styrka'.
    """
    for col_name in row:
        if _norm(col_name) == "styrka":
            v = row[col_name]
            return "" if v is None else str(v).strip()
    return ""


def _strength_sort_key(s: str) -> tuple[float, str]:
    """Sort strengths numerically when possible, lexicographically otherwise.

    '18 mg' sorts before '54 mg' (not by string '18' vs '5').
    Compound strengths like '10 mg/g + 0,25 mg/g' fall back to lex sort.
    """
    if not s:
        return (float("inf"), "")
    m = re.match(r"^\s*([\d.,]+)", s)
    if m:
        try:
            return (float(m.group(1).replace(",", ".")), s.lower())
        except ValueError:
            pass
    return (float("inf"), s.lower())


def _legacy_to_variants(m: dict[str, Any]) -> dict[str, Any]:
    """Promote a pre-variants curated entry to the variants[] shape.

    Older medicines_se.json had one ``npl_id`` and one ``common_forms``
    list per medicine. Map that to a variants array so consumers see a
    uniform shape — one variant per old form, sharing the legacy NPL
    if present, no strength (the legacy schema didn't track strengths).
    """
    out = {k: v for k, v in m.items()
           if k not in ("npl_id", "common_forms")}
    if "variants" in out:
        return out
    variants: list[dict[str, str]] = []
    forms = m.get("common_forms") or []
    legacy_npl = m.get("npl_id")
    if forms:
        for f in forms:
            v: dict[str, str] = {}
            if legacy_npl:
                v["npl_id"] = legacy_npl
            v["form"] = f
            variants.append(v)
    elif legacy_npl:
        variants.append({"npl_id": legacy_npl})
    out["variants"] = variants
    return out


def merge(
    existing: dict[str, Any],
    new_records: Iterable[dict[str, Any]],
) -> tuple[list[dict[str, Any]], dict[str, int]]:
    """Merge new_records into existing['medicines'], preserving aliases.

    Match key is lowercased name. (NPL lookup is no longer 1:1 in the
    variants-aware schema, so the name is the stable join key.)
    Curated aliases survive: unioned with new aliases from the source
    (Tidigare läkemedelsnamn), curated wins on conflict.

    Existing entries that DON'T match anything in the source are
    preserved verbatim — never silently dropped. They may be valid
    medicines that fell out of the dataset (deregistered, renamed,
    misspelled) and the user curated them deliberately. Legacy
    single-npl_id shapes get promoted to variants[] on the way through.
    """
    existing_meds = existing.get("medicines") or []
    by_name = {_norm(m["name"]): m for m in existing_meds}
    matched_existing: set[int] = set()

    merged: list[dict[str, Any]] = []
    seen_keys: set[str] = set()
    s = {
        "added": 0, "updated": 0, "preserved_aliases": 0,
        "unchanged": 0, "merged_aliases": 0, "kept_orphans": 0,
    }

    for rec in new_records:
        name_key = _norm(rec["name"])
        match = by_name.get(name_key)

        if match:
            matched_existing.add(id(match))
            curated = list(match.get("aliases") or [])
            new_aliases = list(rec.get("aliases") or [])
            seen = {_norm(a) for a in curated}
            for a in new_aliases:
                if _norm(a) not in seen:
                    curated.append(a)
                    seen.add(_norm(a))
            name_norm = _norm(rec["name"])
            curated = [a for a in curated if _norm(a) != name_norm]
            curated.sort(key=str.lower)
            if match.get("aliases"):
                s["preserved_aliases"] += 1
            if len(curated) > len(match.get("aliases") or []):
                s["merged_aliases"] += 1
            rec["aliases"] = curated
            # Compare ignoring aliases (and ignoring shape differences
            # for the legacy → variants migration).
            old_compare = {k: v for k, v in _legacy_to_variants(match).items()
                           if k != "aliases"}
            new_compare = {k: v for k, v in rec.items() if k != "aliases"}
            if old_compare != new_compare:
                s["updated"] += 1
            else:
                s["unchanged"] += 1
        else:
            s["added"] += 1

        if name_key in seen_keys:
            continue
        seen_keys.add(name_key)
        merged.append(rec)

    for m in existing_meds:
        if id(m) in matched_existing:
            continue
        copy = _legacy_to_variants(dict(m))
        if copy.get("aliases"):
            name_norm = _norm(copy["name"])
            copy["aliases"] = [
                a for a in copy["aliases"] if _norm(a) != name_norm
            ]
        merged.append(copy)
        s["kept_orphans"] += 1

    merged.sort(key=lambda x: _norm(x["name"]))
    return merged, s


def bump_list_version(existing: dict[str, Any]) -> str:
    today = date.today().strftime("%Y.%m.%d")
    current = existing.get("list_version") or ""
    m = re.match(rf"^{re.escape(today)}-(\d+)$", current)
    if m:
        return f"{today}-{int(m.group(1)) + 1}"
    return f"{today}-1"


def main() -> int:
    parser = argparse.ArgumentParser(
        description=__doc__.split("\n")[0],
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument("--input", "-i", required=True, type=Path,
                        help="Path to Läkemedelsverket export (.xlsx, .csv, "
                             ".tsv, .xml, or .json)")
    parser.add_argument("--output", "-o", required=True, type=Path,
                        help="Path to medicines_se.json")
    parser.add_argument("--dry-run", action="store_true",
                        help="Parse and report stats; do not write output")
    args = parser.parse_args()

    if not args.input.exists():
        print(f"input not found: {args.input}", file=sys.stderr)
        return 2
    if not args.output.exists():
        print(f"output not found (need an existing JSON to merge into): "
              f"{args.output}", file=sys.stderr)
        return 2

    existing = json.loads(args.output.read_text(encoding="utf-8"))
    raw_records = load_records(args.input)
    print(f"loaded {len(raw_records)} raw rows from {args.input.name}")

    records, gstats = group_and_normalize(raw_records)
    print(
        f"filter+group: kept {gstats['kept_rows']} rows / "
        f"{gstats['total_rows']} total "
        f"(skipped: {gstats['skipped_status']} non-current status, "
        f"{gstats['skipped_vet']} veterinary, "
        f"{gstats['skipped_no_name']} missing name, "
        f"{gstats['dedup_parallel_imports']} parallel-import dupes folded)"
    )
    print(
        f"  → {gstats['unique_medicines']} unique medicines, "
        f"{gstats['total_variants']} total variants"
    )

    merged, mstats = merge(existing, records)
    print(
        f"merge: +{mstats['added']} added, "
        f"~{mstats['updated']} updated, "
        f"={mstats['unchanged']} unchanged, "
        f"aliases preserved on {mstats['preserved_aliases']} entries, "
        f"alias lists grew on {mstats['merged_aliases']} entries"
    )
    if mstats["kept_orphans"]:
        print(
            f"  ⚠ kept {mstats['kept_orphans']} curated entries that "
            "weren't in the source (deregistered, renamed, veterinary, "
            "or otherwise out-of-band — review and prune as needed)"
        )

    new_version = bump_list_version(existing)
    print(f"list_version: {existing.get('list_version')} → {new_version}")

    if args.dry_run:
        print("(dry-run — output unchanged)")
        return 0

    out = dict(existing)
    out["schema_version"] = 2  # variants[] replaces npl_id+common_forms
    out["list_version"] = new_version
    out["updated"] = date.today().isoformat()
    out["medicines"] = merged
    out["fields"] = {
        "name": "Display name (usually brand). Required.",
        "aliases": (
            "Alternative spellings, generic names, common misspellings, "
            "former product names. Searched alongside name in the dropdown."
        ),
        "active_substance": "Active substance (Swedish or Latin name).",
        "atc_code": (
            "WHO ATC code, where confidently known. Empty string if "
            "uncertain."
        ),
        "variants": (
            "List of {npl_id, strength, form} — one per distinct "
            "strength+form combination. Enables per-variant deep-linking "
            "to Fass / Läkemedelsverket and strength/form picker UI."
        ),
    }
    out["notice"] = (
        "Compiled from Läkemedelsverket's open-data register "
        "(Sök läkemedelsfakta, Sveriges dataportal dataset 140_5467) "
        "with curated aliases added by contributors. NOT a clinical "
        "reference — verify ATC codes and active substances against "
        "authoritative sources (Sök VARA, FASS, Läkemedelsboken) before "
        "relying on them. Aliases are former product names (from the "
        "Tidigare-läkemedelsnamn column) plus common misspellings, "
        "generic names, and alternate brand names added manually to "
        "help search find the right medicine. Contributions welcome "
        "via PR — see README."
    )

    args.output.write_text(
        json.dumps(out, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    print(f"wrote {args.output} ({len(merged)} medicines)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
